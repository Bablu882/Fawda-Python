from django.shortcuts import render
from authentication.models import User
from booking.models import JobBooking
from django.http import JsonResponse
from jobs.models import JobMachine,JobSahayak
from rest_framework.response import Response
from .models import BookingHistoryMachine,BookingHistorySahayak,ClientInformation
import uuid
import openpyxl
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny,IsAuthenticated
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from .serializers import *
from authentication.views import BearerTokenAuthentication
from django.db.models import Max
from datetime import datetime
from rest_framework.pagination import PageNumberPagination
from rest_framework.authentication import SessionAuthentication
from rest_framework.permissions import IsAdminUser


class JobDetailsAdmin(APIView):
    authentication_classes=[SessionAuthentication]
    permission_classes=[IsAdminUser,]
    PAGE_SIZE=10
    def get(self,request,format=None):
        status = request.GET.get('status')
        # print(status)
        data=[]
        if status == 'Pending' or status == 'Timeout':
            job_sahayak=JobSahayak.objects.filter(status=status).order_by('-id')
            job_machine=JobMachine.objects.filter(status=status).order_by('-id')
            for jobs in job_sahayak:         
                data.append({
                    'id':jobs.id,
                    'job_number':jobs.job_number,
                    'Job_type':jobs.job_type,
                    'Job_posting_date':jobs.date,
                    'Job_booking_date':'Null',
                    'Job_status':jobs.status,
                    'payment_to_service_provider':jobs.payment_your,
                    'mobile_no_for_payment':'Null',
                    'upi_id_for_payment':'Null',
                    'Payment_status':'Null'
                })
            for jobs in job_machine:
                data.append({
                    'id':jobs.id,
                    'job_number':jobs.job_number,
                    'Job_type':jobs.job_type,
                    'Job_posting_date':jobs.date,
                    'Job_booking_date':'Null',
                    'Job_status':jobs.status,
                    'payment_to_service_provider':jobs.payment_your,
                    'mobile_no_for_payment':'Null',
                    'upi_id_for_payment':'Null',
                    'Payment_status':'Null'

                })
        else:
            job_booking=JobBooking.objects.filter(status=status,is_admin_paid='Pending').order_by('-id')
            for jobs in job_booking:
                if jobs.jobsahayak:
                    if jobs.jobsahayak.job_type == 'theke_pe_kam':
                        data.append({
                            'id':jobs.id,
                            'job_number':jobs.jobsahayak.job_number,
                            'Job_type':jobs.jobsahayak.job_type,
                            'Job_posting_date':jobs.jobsahayak.date,
                            'Job_booking_date':jobs.date_booked,
                            'Job_status':jobs.status,
                            'payment_to_service_provider':jobs.payment_your,
                            'mobile_no_for_payment':jobs.booking_user.mobile_no,
                            'upi_id_for_payment':jobs.booking_user.profile.upiid,
                            'Payment_status':jobs.is_admin_paid

                        })
                    else:
                        data.append({
                            'id':jobs.id,
                            'job_number':jobs.jobsahayak.job_number,
                            'Job_type':jobs.jobsahayak.job_type,
                            'Job_posting_date':jobs.jobsahayak.date,
                            'Job_booking_date':jobs.date_booked,
                            'Job_status':jobs.status,
                            'payment_to_service_provider':jobs.payment_your,
                            'mobile_no_for_payment':jobs.booking_user.mobile_no,
                            'upi_id_for_payment':jobs.booking_user.profile.upiid,
                            'Payment_status':jobs.is_admin_paid

                        })
                else:
                    data.append({
                        'id':jobs.id,
                        'job_number':jobs.jobmachine.job_number,
                        'Job_type':jobs.jobmachine.job_type,
                        'Job_posting_date':jobs.jobmachine.date,
                        'Job_booking_date':jobs.date_booked,
                        'Job_status':jobs.status,
                        'payment_to_service_provider':jobs.payment_your,
                        'mobile_no_for_payment':jobs.booking_user.mobile_no,
                        'upi_id_for_payment':jobs.booking_user.profile.upiid,
                        'Payment_status':jobs.is_admin_paid


                    })    
        paginator = PageNumberPagination()
        paginator.page_size = self.PAGE_SIZE
        paginated_result = paginator.paginate_queryset(data, request)
        response_data = {'page': paginator.page.number, 'total_pages': paginator.page.paginator.num_pages, 'results': paginated_result}
    # Add next page URL to response
        if paginator.page.has_next():
            base_url = request.build_absolute_uri().split('?')[0]
            response_data['next'] = f"{base_url}?page={paginator.page.next_page_number()}"            
        return Response({'data':response_data})



class JobDetailsAdminPanel(APIView):
    authentication_classes=[SessionAuthentication]
    permission_classes=[IsAdminUser,]
    def get(self,request,format=None):
        data=request.GET.get('id')
        job_number=request.GET.get('job_number')
        data_list=[]
        if JobSahayak.objects.filter(id=data,job_number=job_number).exists():
            details1=JobSahayak.objects.get(id=data)
            data_list.append({
                'grahak_name':details1.grahak.profile.name,
                'grahak_gender':details1.grahak.profile.name,
                'grahak_phone':details1.grahak.mobile_no,
                'grahak_mohalla':details1.grahak.profile.mohalla,
                'grahak_village':details1.grahak.profile.village,
                'grahak_state':details1.grahak.profile.state,
                'grahak_district':details1.grahak.profile.district,
                'grahak_pincode':details1.grahak.profile.pincode,
                'grahak_age':details1.grahak.profile.age,
                'grahak_upiid':details1.grahak.profile.upiid,
                'status':details1.status,
                'desc':details1.description,
                 
                'heading':"Sahayak Details",
                'name':"None",
                'gender':"None",
                'phone':"None",
                'mohalla':"None",
                'village':"None",
                'state':"None",
                'district':"None",
                'pincode':"None",
                'age':"None",
                'upiid':"None",

            })
        elif JobMachine.objects.filter(id=data,job_number=job_number).exists():
            details2=JobMachine.objects.get(id=data)
            data_list.append({
                'grahak_name':details2.grahak.profile.name,
                'grahak_gender':details2.grahak.profile.gender,
                'grahak_phone':details2.grahak.mobile_no,
                'grahak_mohalla':details2.grahak.profile.mohalla,
                'grahak_village':details2.grahak.profile.village,
                'grahak_state':details2.grahak.profile.state,
                'grahak_district':details2.grahak.profile.district,
                'grahak_pincode':details2.grahak.profile.pincode,
                'grahak_age':details2.grahak.profile.age,
                'grahak_upiid':details2.grahak.profile.upiid,
                'status':details2.status,
                'desc':(details2.work_type) + " , "  + (details2.machine) + "\n" + (details2.description),
                
                'heading':"Sahayak Details",
                'name':"None",
                'gender':"None",
                'phone':"None",
                'mohalla':"None",
                'village':"None",
                'state':"None",
                'district':"None",
                'pincode':"None",
                'age':"None",
                'upiid':"None",

            })
        elif JobBooking.objects.filter(id=data).exists():
            details3=JobBooking.objects.get(id=data)
            if details3.jobsahayak:
                data_list.append({
                    'grahak_name':details3.jobsahayak.grahak.profile.name,
                    'grahak_gender':details3.jobsahayak.grahak.profile.gender,
                    'grahak_phone':details3.jobsahayak.grahak.mobile_no,
                    'grahak_mohalla':details3.jobsahayak.grahak.profile.mohalla,
                    'grahak_village':details3.jobsahayak.grahak.profile.village,
                    'grahak_state':details3.jobsahayak.grahak.profile.state,
                    'grahak_district':details3.jobsahayak.grahak.profile.district,
                    'grahak_pincode':details3.jobsahayak.grahak.profile.pincode,
                    'grahak_age':details3.jobsahayak.grahak.profile.age,
                    'grahak_upiid':details3.jobsahayak.grahak.profile.upiid,
                    'desc':details3.jobsahayak.description,
                    'status':details3.status,
                    
                    'heading':"Sahayak Details",
                    'name':details3.booking_user.profile.name,
                    'gender':details3.booking_user.profile.gender,
                    'phone':details3.booking_user.mobile_no,
                    'mohalla':details3.booking_user.profile.mohalla,
                    'village':details3.booking_user.profile.village,
                    'state':details3.booking_user.profile.state,
                    'district':details3.booking_user.profile.district,
                    'pincode':details3.booking_user.profile.pincode,
                    'age':details3.booking_user.profile.age,
                    'upiid':details3.booking_user.profile.upiid,

                })
            else:
                data_list.append({
                    'grahak_name':details3.jobmachine.grahak.profile.name,
                    'grahak_gender':details3.jobmachine.grahak.profile.gender,
                    'grahak_phone':details3.jobmachine.grahak.mobile_no,
                    'grahak_mohalla':details3.jobmachine.grahak.profile.mohalla,
                    'grahak_village':details3.jobmachine.grahak.profile.village,
                    'grahak_state':details3.jobmachine.grahak.profile.state,
                    'grahak_district':details3.jobmachine.grahak.profile.district,
                    'grahak_pincode':details3.jobmachine.grahak.profile.pincode,
                    'grahak_age':details3.jobmachine.grahak.profile.age,
                    'grahak_upiid':details3.jobmachine.grahak.profile.upiid,
                    'desc':details3.jobmachine.work_type + " , " + (details3.jobmachine.machine) + "\n" + (details3.jobmachine.description),
                    'status':details3.status,
                    
                    'heading':"Machine Malik Details",
                    'name':details3.booking_user.profile.name,
                    'gender':details3.booking_user.profile.gender,
                    'phone':details3.booking_user.mobile_no,
                    'mohalla':details3.booking_user.profile.mohalla,
                    'village':details3.booking_user.profile.village,
                    'state':details3.booking_user.profile.state,
                    'district':details3.booking_user.profile.district,
                    'pincode':details3.booking_user.profile.pincode,
                    'age':details3.booking_user.profile.age,
                    'upiid':details3.booking_user.profile.upiid,
                })  

        else:
            return Response({'Job not exists !'})         
        return Response(data_list)    
    

class AdminPaymentStatus(APIView):
    authentication_classes=[SessionAuthentication,]
    permission_classes=[IsAdminUser,]
    def post(self,request,format=None):
        get_id=request.POST.get('id')
        # if request.user.is_superuser:
        if JobBooking.objects.filter(pk=get_id).exists():
            get=JobBooking.objects.get(pk=get_id)
            if get.status == 'Completed':
                get.is_admin_paid='Paid'
                get.save()
                if get.jobsahayak:
                    booking=BookingHistorySahayak.objects.create(
                        grahak_name=get.jobsahayak.grahak.profile.name,
                        grahak_mobile_no=get.jobsahayak.grahak.mobile_no,
                        job_type=get.jobsahayak.job_type,
                        job_number=get.jobsahayak.job_number,
                        job_posting_date=get.jobsahayak.date,
                        job_booking_date=get.date_booked,
                        job_status=get.jobsahayak.status,
                        payment_status_by_admin=get.is_admin_paid,
                        paid_to_service_provider=get.payment_your,
                        paid_by_grahak=get.total_amount,
                        sahayak_name=get.booking_user.profile.name,
                        sahayak_mobile_no=get.booking_user.mobile_no,
                        booking_id=get.id
                    )
                else:
                    booking=BookingHistoryMachine.objects.create(
                        grahak_name=get.jobmachine.grahak.profile.name,
                        grahak_mobile_no=get.jobmachine.grahak.mobile_no,
                        job_type=get.jobmachine.job_type,
                        job_number=get.jobmachine.job_number,
                        job_posting_date=get.jobmachine.date,
                        job_booking_date=get.date_booked,
                        job_status=get.jobmachine.status,
                        payment_status_by_admin=get.is_admin_paid,
                        paid_to_service_provider=get.payment_your,
                        paid_by_grahak=get.total_amount,
                        machine_malik_name=get.booking_user.profile.name,
                        machine_malik_mobile_no=get.booking_user.mobile_no,
                        booking_id=get.id
                    )  
            elif get.status == 'Cancelled-After-Payment' or get.status == 'Rejected-After-Payment':
                get.is_admin_paid ='Refunded'
                get.status = 'Admin-Refunded'
                get.save()
                            
        return Response({'msg':{'Payment status updated successfully !'}})    


def booking_history_sahayak(request):
    bookings = BookingHistorySahayak.objects.all()
    return render(request, 'booking_history.html', {'bookings': bookings})


@staff_member_required
def my_custom_view(request):
    bookings=BookingHistorySahayak.objects.all()
    bookings2=BookingHistoryMachine.objects.all()
    return render(request, 'admin/custom_home.html', {'bookings':bookings,'bookings2':bookings2})

@staff_member_required
def booking_log_history(request):
    bookings=JobBooking.history.all()
    print(bookings)
    return render(request, 'admin/booking_log.html', {'historys':bookings})


@staff_member_required
def custom_users_view(request):
    user_list=[]
    users=User.objects.all()
    for user in users:
        user_list.append({
            'name':user.profile.name,
            'address':user.profile.village + ", " + user.profile.mohalla + ", " + user.profile.district + ", " + user.profile.state,
            'gender':user.profile.gender,
            'phone':user.mobile_no,
            'user_type':user.user_type
        })
    return render(request, 'admin/user_history.html', {'users':user_list})


def export_users_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="users.xlsx"'

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    worksheet = workbook.active

    # Add column headings
    columns = ['Name', 'Address', 'Gender', 'Phone', 'User Type']
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Add data rows
    users = User.objects.all()
    for user in users:
        row_num += 1
        row = [
            user.profile.name,
            f"{user.profile.village}, {user.profile.mohalla}, {user.profile.district}, {user.profile.state}",
            user.profile.gender,
            user.mobile_no,
            user.user_type,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the Excel file
    workbook.save(response)

    return response


def export_booking_history_sahayak_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="booking_history.xlsx"'

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    worksheet = workbook.active

    # Add column headings
    columns = [
        'Grahak Name',
        'Grahak Mobile No',
        'Job Type',
        'Job Number',
        'Job Posting Date',
        'Job Booking Date',
        'Job Status',
        'Payment Status By Admin',
        'Paid To Service Provider',
        'Paid By Grahak',
        'Sahayak Name',
        'Sahayak Mobile No',
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Add data rows
    booking_history = BookingHistorySahayak.objects.all()
    for booking in booking_history:
        row_num += 1
        row = [
            booking.grahak_name,
            booking.grahak_mobile_no,
            booking.job_type,
            booking.job_number,
            booking.job_posting_date,
            booking.job_booking_date,
            booking.job_status,
            booking.payment_status_by_admin,
            booking.paid_to_service_provider,
            booking.paid_by_grahak,
            booking.sahayak_name,
            booking.sahayak_mobile_no,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the Excel file
    workbook.save(response)

    return response



def export_booking_history_machine_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="booking_history_machine.xlsx"'

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    worksheet = workbook.active

    # Add column headings
    columns = [
        'Grahak Name',
        'Grahak Mobile No',
        'Job Type',
        'Job Number',
        'Job Posting Date',
        'Job Booking Date',
        'Job Status',
        'Payment Status By Admin',
        'Paid To Service Provider',
        'Paid By Grahak',
        'Machine Malik Name',
        'Machine Malik Mobile No',
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Add data rows
    booking_history = BookingHistoryMachine.objects.all()
    for booking in booking_history:
        row_num += 1
        row = [
            booking.grahak_name,
            booking.grahak_mobile_no,
            booking.job_type,
            booking.job_number,
            booking.job_posting_date,
            booking.job_booking_date,
            booking.job_status,
            booking.payment_status_by_admin,
            booking.paid_to_service_provider,
            booking.paid_by_grahak,
            booking.machine_malik_name,
            booking.machine_malik_mobile_no,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the Excel file
    workbook.save(response)

    return response


class TermsAndCondition(APIView):
    permission_classes=[AllowAny,]
    def get(self,request,format=None):
        terms=ClientInformation.objects.all()
        serializer=ClientInformationsSerializer(terms,many=True)
        return Response(serializer.data)

class ClientUserInfo(APIView):
    permission_classes=[IsAuthenticated,]
    authentication_classes=[BearerTokenAuthentication,]
    def get(self,request,format=None):
        app_version= AppVersion.objects.order_by('-release_date').first()
        latest_info = ClientInformation.objects.aggregate(Max('created_at'))
        if latest_info['created_at__max']:
            latest_info = ClientInformation.objects.get(created_at=latest_info['created_at__max'])
            client_info = {
                'privacy_policy': latest_info.privacy_policy,
                'terms_condition': latest_info.terms_condition,
                'about_us':latest_info.about_us,
                'phone_no': latest_info.phone_no,
                
            }
        user_info=({
            'phone':request.user.mobile_no,
            'name':request.user.profile.name,
            'mohalla':request.user.profile.mohalla,
            'village':request.user.profile.village,
            'gender':request.user.profile.gender,
            'state':request.user.profile.state,
            'district':request.user.profile.district,
            'pincode':request.user.profile.pincode,
        })
        return Response(
            {'client_info': client_info, 'app_version': str(app_version), 'user_details': user_info}
        )
    



def BookingInvoiceView(request):
    # if  request.user.is_superuser:
    invoice_data=[]
    bookings=JobBooking.objects.all().filter(is_admin_paid='Paid')
    latest_info = ClientInformation.objects.aggregate(Max('created_at'))
    if latest_info['created_at__max']:
        latest_info = ClientInformation.objects.get(created_at=latest_info['created_at__max'])  
    for booking in bookings:
        if booking.jobsahayak:
            invoice_data.append({
                'grahak_name':booking.jobsahayak.grahak.profile.name,
                'job_number':booking.jobsahayak.job_number,
                'grahak_address':f"{booking.jobsahayak.grahak.profile.village},{booking.jobsahayak.grahak.profile.mohalla},{booking.jobsahayak.grahak.profile.district},{booking.jobsahayak.grahak.profile.state}",
                'date_invoice_generation':'date',
                'fawda_fee':booking.fawda_fee *2,
                'busigness_gst':latest_info.gst_no,
                'busigness_name':latest_info.business_name,
                'client_address':latest_info.client_address,
                'sahayak_name':booking.booking_user.profile.name,
                'sahayak_address':f"{booking.booking_user.profile.village},{booking.booking_user.profile.mohalla},{booking.booking_user.profile.district},{booking.booking_user.profile.state}",
                'payment':booking.payment_your
            })
        else:
            invoice_data.append({
                'grahak_name':booking.jobmachine.grahak.profile.name,
                'job_number':booking.jobmachine.job_number,
                'grahak_address':f"{booking.jobmachine.grahak.profile.village},{booking.jobmachine.grahak.profile.mohalla},{booking.jobmachine.grahak.profile.district},{booking.jobmachine.grahak.profile.state}",
                'date_invoice_generation':'date',
                'fawda_fee':booking.fawda_fee *2,
                'busigness_gst':latest_info.gst_no,
                'busigness_name':latest_info.business_name,
                'client_address':latest_info.client_address,
                'machine_malik_name':booking.booking_user.profile.name,
                'machine_malik_address':f"{booking.booking_user.profile.village},{booking.booking_user.profile.mohalla},{booking.booking_user.profile.district},{booking.booking_user.profile.state}",
                'payment':booking.payment_your
            })
    return render(request,'admin/custom_home.html',context=invoice_data)        



def BookingInvoice(request,id):
    print(id)
    if JobBooking.objects.filter(pk=id,is_admin_paid='Paid'):
        booking=JobBooking.objects.get(pk=id)
        latest_info = ClientInformation.objects.aggregate(Max('created_at'))
        invoice_data = None
        if latest_info['created_at__max']:
            latest_info = ClientInformation.objects.get(created_at=latest_info['created_at__max'])  
        if booking.jobsahayak:
            invoice_data=({
                'grahak_name':booking.jobsahayak.grahak.profile.name,
                'job_number':booking.jobsahayak.job_number,
                'grahak_address':f"{booking.jobsahayak.grahak.profile.village},{booking.jobsahayak.grahak.profile.mohalla},{booking.jobsahayak.grahak.profile.district},{booking.jobsahayak.grahak.profile.state}",
                'date_invoice_generation':'date',
                'fawda_fee':(float(booking.fawda_fee_grahak) + float(booking.fawda_fee_sahayak)),
                'busigness_gst':latest_info.gst_no,
                'busigness_name':latest_info.business_name,
                'client_address':latest_info.client_address,
                'sahayak_name':booking.booking_user.profile.name,
                'sahayak_address':f"{booking.booking_user.profile.village},{booking.booking_user.profile.mohalla},{booking.booking_user.profile.district},{booking.booking_user.profile.state}",
                'payment':booking.payment_your
            })
        else:
            invoice_data=({
                'grahak_name':booking.jobmachine.grahak.profile.name,
                'job_number':booking.jobmachine.job_number,
                'grahak_address':f"{booking.jobmachine.grahak.profile.village},{booking.jobmachine.grahak.profile.mohalla},{booking.jobmachine.grahak.profile.district},{booking.jobmachine.grahak.profile.state}",
                'date_invoice_generation':'date',
                'fawda_fee':(float(booking.fawda_fee_grahak) + float(booking.fawda_fee_machine)),
                'busigness_gst':latest_info.gst_no,
                'busigness_name':latest_info.business_name,
                'client_address':latest_info.client_address,
                'machine_malik_name':booking.booking_user.profile.name,
                'machine_malik_address':f"{booking.booking_user.profile.village},{booking.booking_user.profile.mohalla},{booking.booking_user.profile.district},{booking.booking_user.profile.state}",
                'payment':booking.payment_your
            })
        print(invoice_data)    
    return render(request,'admin/custom_home.html',context={'invoice_data':invoice_data})        




class BookingInvoiceApiView(APIView):
    authentication_classes=[SessionAuthentication]
    permission_classes=[IsAdminUser,]
    def get(self,request,format=None):
        booking_id=request.GET.get('id')
        invoice_data = None
        if JobBooking.objects.filter(pk=booking_id,is_admin_paid='Paid'):
            booking=JobBooking.objects.get(pk=booking_id)
            latest_info = ClientInformation.objects.aggregate(Max('created_at'))
            if latest_info['created_at__max']:
                latest_info = ClientInformation.objects.get(created_at=latest_info['created_at__max'])  
            if booking.jobsahayak:
                invoice_data=({
                    'grahak_name':booking.jobsahayak.grahak.profile.name,
                    'job_number':booking.jobsahayak.job_number,
                    'grahak_address':f"{booking.jobsahayak.grahak.profile.village},{booking.jobsahayak.grahak.profile.mohalla},{booking.jobsahayak.grahak.profile.district},{booking.jobsahayak.grahak.profile.state}",
                    'date_invoice_generation':datetime.now().strftime('%Y-%m-%d'),
                    'fawda_fee':(float(booking.fawda_fee_grahak) + float(booking.fawda_fee_sahayak)),
                    'busigness_gst':latest_info.gst_no,
                    'busigness_name':latest_info.business_name,
                    'client_address':latest_info.client_address,
                    'provider_name':booking.booking_user.profile.name,
                    'provider_address':f"{booking.booking_user.profile.village},{booking.booking_user.profile.mohalla},{booking.booking_user.profile.district},{booking.booking_user.profile.state}",
                    'payment':booking.payment_your
                })
            else:
                invoice_data=({
                    'grahak_name':booking.jobmachine.grahak.profile.name,
                    'job_number':booking.jobmachine.job_number,
                    'grahak_address':f"{booking.jobmachine.grahak.profile.village},{booking.jobmachine.grahak.profile.mohalla},{booking.jobmachine.grahak.profile.district},{booking.jobmachine.grahak.profile.state}",
                    'date_invoice_generation':datetime.now().strftime('%Y-%m-%d'),
                    'fawda_fee':(float(booking.fawda_fee_grahak) + float(booking.fawda_fee_machine)),
                    'busigness_gst':latest_info.gst_no,
                    'busigness_name':latest_info.business_name,
                    'client_address':latest_info.client_address,
                    'provider_name':booking.booking_user.profile.name,
                    'provider_address':f"{booking.booking_user.profile.village},{booking.booking_user.profile.mohalla},{booking.booking_user.profile.district},{booking.booking_user.profile.state}",
                    'payment':booking.payment_your
                })
            print(invoice_data)
            print(booking.fawda_fee)    
        return Response({'invoice_data':invoice_data})        


##--------------------------------expo nortifications------------------------------------------###


# import json
# import requests
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from django.conf import settings

# # Import the expo-server-sdk-python library
# from expo_server_sdk import PushClient
# from expo_server_sdk import PushMessage
# from expo_server_sdk import ExpoPushMessage
# from expo_server_sdk import PushResponseError
# from expo_server_sdk import PushServerError

# class PushNotificationView(APIView):
#     authentication_classes=[IsAuthenticated,]
#     permission_classes=[BearerTokenAuthentication,]
#     def post(self, request, format=None):
#         # Get the user ID from the request data
#         user_id = request.user.id
        
#         # Get the user's expoPushToken from the database
#         user = User.objects.get(id=user_id)
#         expo_push_token = user.expo_push_token
        
#         # Create the notification message
#         message = PushMessage(
#             to=expo_push_token,
#             title="New Message",
#             body="You have a new message"
#         )
        
#         # Send the notification using the expo-server-sdk-python package
#         try:
#             response = PushClient().publish([message])
#             print(response)
#             return Response({"status": "success"})
#         except PushResponseError as exc:
#             print(exc.errors)
#             return Response({"status": "error", "message": "Push notification error"})
#         except (PushServerError, Exception) as exc:
#             print(exc)
#             return Response({"status": "error", "message": "Push notification error"})
